from datetime import datetime, date
from dataclasses import dataclass

from openpyxl import load_workbook, Workbook

WORKBOOK_PATH = "CoreMark.xlsx"
OUTPUT_PATH = "output.xlsx"

YEARS = [2024, 2025, 2026]
DEFAULT_START = date(2024, 10, 23)
DEFAULT_END = date(2026, 12, 31)

REQUIRED_COLUMNS = {
    "ITEM CODE",
    "UNIT TYPE",
    "DESCRIPTION",
    "VALUE",
    "DATES"
}

@dataclass
class Record:
    item_code: str
    unit_type: str
    description: str
    value: str
    start_date: date
    end_date: date

########################################
### INGEST
########################################
def validate_headers(header_map):
    missing = REQUIRED_COLUMNS - header_map.keys()

    if missing:
        raise ValueError(f"Missing required columns: {missing}")


def build_header_map(ws):
    header_map = {}

    for idx, cell in enumerate(ws[1]):
        if cell.value is None:
            continue
        header_map[cell.value.strip().upper()] = idx

    return header_map


def parse_dates(value):
    if value is None or str(value).strip() == "":
        return DEFAULT_START, DEFAULT_END

    value = str(value).strip()

    if " to " in value:
        start, end = value.split(" to ")
        return (
            datetime.strptime(start, "%m/%d/%Y").date(),
            datetime.strptime(end, "%m/%d/%Y").date(),
        )

    if " and later" in value:
        start = value.replace(" and later", "")
        return datetime.strptime(start, "%m/%d/%Y").date(), DEFAULT_END

    if value.startswith("On or before "):
        end = value.replace("On or before ", "")
        return DEFAULT_START, datetime.strptime(end, "%m/%d/%Y").date()

    raise ValueError(value)


def load_records(path):
    wb = load_workbook(path)
    ws = wb["Sheet1"]

    header_map = build_header_map(ws)
    validate_headers(header_map)

    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        start, end = parse_dates(row[header_map["DATES"]])

        record = Record(
            item_code=row[header_map["ITEM CODE"]],
            unit_type=row[header_map["UNIT TYPE"]],
            description=row[header_map["DESCRIPTION"]],
            value=row[header_map["VALUE"]],
            start_date=start,
            end_date=end
        )

        records.append(record)

    return records


########################################
### EMIT
########################################
def overlaps_year(start, end, year):
    return not (
        end < date(year, 1, 1) or
        start > date(year, 12, 31)
    )


def bound_to_year(start, end, year):
    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)
    return max(start, year_start), min(end, year_end)


def emit_workbook(records, path):

    wb = Workbook()
    sheets = {}

    for sheetname in wb.sheetnames:
        del wb[sheetname]

    for year in YEARS:
        ws = wb.create_sheet(str(year))
        sheets[year] = ws
        ws.append(["ITEM CODE", "UNIT TYPE", "DESCRIPTION", "VALUE", "START DATE", "END DATE"])

    for r in records:
        for year in YEARS:
            if overlaps_year(r.start_date, r.end_date, year):
                start, end = bound_to_year(r.start_date, r.end_date, year)
                sheets[year].append([
                    r.item_code,
                    r.unit_type,
                    r.description,
                    r.value,
                    start,
                    end
                ])

    wb.save(path)


########################################
### MAIN
########################################

if __name__ == "__main__":
    records = load_records(WORKBOOK_PATH)
    emit_workbook(records, OUTPUT_PATH)