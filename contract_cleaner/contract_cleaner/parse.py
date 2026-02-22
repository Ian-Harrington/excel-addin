from decimal import Decimal
import logging

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from contract_cleaner.models import LineItem, HeaderCellMapping, SheetFormat


log = logging.getLogger(__name__)


USE_DEFAULT_SHEET_FORMAT = False

ITEM_NUMBER_HEADERS = ["ID", "Item Number", "Identifier", "MVC", "SKU", "Part Number", "Item #", "Item No.", "Item Code", "Code"]
DESCRIPTION_HEADERS = ["Desc", "Description"]
PRICE_HEADERS = ["Cost", "Price", "Amount"]

HEADER_SEARCH_RANGE = "A1:Z20"



def parse_workbook(wb: Workbook) -> list[SheetFormat]:
    """Determine sheets to parse and their header mappings"""
    if len(wb.sheetnames) == 0:
        raise ValueError("Workbook has no sheets")
    
    sheet_formats = []
    if USE_DEFAULT_SHEET_FORMAT:
        log.info("Using default sheet format for 'Items' sheet")
        if "Items" in wb.sheetnames:
            return [SheetFormat("Items", HeaderCellMapping())]
    else:
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            format = discover_sheet_format(sheet)
            sheet_formats.append(format) if format else log.info(f"No header row found for sheet '{sheetname}', skipping")
    
    if len(sheet_formats) == 0:
        raise ValueError("No sheets with valid header row found")
    return sheet_formats


def discover_sheet_format(sheet: Worksheet) -> SheetFormat | None:
    """Searches for header row and determines column mapping based on header names"""
    for row in sheet[HEADER_SEARCH_RANGE]:
        identifier_col = None
        description_col = None
        price_col = None
        for cell in row:
            if cell.value is None:
                continue
            cell_value = str(cell.value).strip().lower()
            if cell_value in [val.lower() for val in ITEM_NUMBER_HEADERS]:
                identifier_col = cell.column
            elif cell_value in [val.lower() for val in DESCRIPTION_HEADERS]:
                description_col = cell.column
            elif cell_value in [val.lower() for val in PRICE_HEADERS]:
                price_col = cell.column
        if identifier_col and description_col and price_col:
            return SheetFormat(
                name=sheet.title,
                header_mapping=HeaderCellMapping(
                    row=cell.row,
                    identifier=identifier_col,
                    description=description_col,
                    price=price_col,
                )
            )
    return None


def parse_sheet(sheet: Worksheet, format: HeaderCellMapping) -> list[LineItem]:
    items: list[LineItem] = []
    max_col = max(format.identifier, format.description, format.price)
    for row in sheet.iter_rows(min_row=format.row, max_row=999, max_col=max_col, values_only=True):
        print(f"Parsing row: {row}")
        item = _parse_row(row, format)
        items.append(item)
    return items


def _parse_row(row: tuple, format: HeaderCellMapping) -> LineItem:
    identifier = _parse_string(row[format.identifier - 1])
    description = _parse_string(row[format.description - 1])
    price = _parse_dollars(row[format.price - 1])

    return LineItem(identifier, description, price)


def _parse_string(value: str | None) -> str | None:
    if value is None:
        return None
    stripped_value = str(value).strip()
    return stripped_value if stripped_value else None


def _parse_dollars(value: str | None) -> Decimal | None:
    if value is None:
        return None
    try:
        cleaned_value = str(value).replace("$", "").replace(",", "").strip()
        return Decimal(cleaned_value) if cleaned_value else None
    except Exception as e:
        log.warning(f"Error parsing decimal value '{value}': {e}")
        return None