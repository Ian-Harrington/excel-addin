from pathlib import Path

from openpyxl import load_workbook, Workbook

from contract_cleaner.models import ParsedSheet


def write_items_to_csv(output: list[ParsedSheet], output_path: Path) -> None:
    for sheet in output:
        csv_output_path = output_path.parent / f"{output_path.stem}_{sheet.name}.csv"
        with open(csv_output_path, "w") as f:
            for item in sheet.items:
                f.write(",".join([str(x) for x in item.as_tuple()]) + "\n")


def write_items_to_workbook(output: list[ParsedSheet], output_path: Path) -> None:
    wb = Workbook()
    
    # remove default sheet(s)
    for sheetname in wb.sheetnames:
        del wb[sheetname]
    
    for sheet_result in output:
        wb.create_sheet(title=sheet_result.name)
        sheet = wb[sheet_result.name]
        for item in sheet_result.items:
            sheet.append(item.as_tuple())

    wb.save(output_path)


def load_workbook_from_path(path: Path) -> Workbook:
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    return load_workbook(path)